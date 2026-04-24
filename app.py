from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import psycopg2
import psycopg2.extras
import os
import time
import json
import io

app = Flask(__name__)
CORS(app)

# ============================================================
# CACHE SIMPLES EM MEMÓRIA
# Guarda resultados por 5 minutos para não bater no banco
# toda vez que alguém acessa a página
# ============================================================
_cache = {}
CACHE_TTL = 28800  # 8 horas em segundos

def cache_get(key):
    if key in _cache:
        valor, timestamp = _cache[key]
        if time.time() - timestamp < CACHE_TTL:
            return valor
    return None

def cache_set(key, valor):
    _cache[key] = (valor, time.time())

def cache_clear():
    _cache.clear()

# ============================================================
# CONEXÃO COM SUPABASE
# ============================================================
def get_conn():
    return psycopg2.connect(
        host     = os.environ.get('DB_HOST'),
        port     = int(os.environ.get('DB_PORT', 6543)),
        database = os.environ.get('DB_NAME', 'postgres'),
        user     = os.environ.get('DB_USER'),
        password = os.environ.get('DB_PASS'),
        sslmode  = 'require',
        connect_timeout = 10,
    )

def _serializar_row(row):
    import datetime
    out = {}
    for k, v in row.items():
        if isinstance(v, datetime.date):
            out[k] = v.isoformat()   # garante string "YYYY-MM-DD" no JSON
        else:
            out[k] = v
    return out

def consultar(sql, params=()):
    conn   = get_conn()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cursor.execute(sql, params)
    resultado = cursor.fetchall()
    cursor.close()
    conn.close()
    return [_serializar_row(dict(row)) for row in resultado]

# ============================================================
# MONTA FILTROS
# ============================================================
def montar_filtros(args):
    condicoes = []
    params    = []

    anos = args.getlist('ano')
    if anos:
        placeholders = ','.join(['%s'] * len(anos))
        condicoes.append(f"ano IN ({placeholders})")
        params.extend([int(a) for a in anos])

    meses = args.getlist('mes')
    if meses:
        placeholders = ','.join(['%s'] * len(meses))
        condicoes.append(f"mes IN ({placeholders})")
        params.extend([int(m) for m in meses])

    unidade = args.get('unidade')
    if unidade:
        condicoes.append("unidade = %s")
        params.append(unidade)

    uf = args.get('uf')
    if uf:
        condicoes.append("uf = %s")
        params.append(uf)

    tipo = args.get('tipo')
    if tipo:
        condicoes.append("tipo_operacao = %s")
        params.append(tipo)

    marca = args.get('marca')
    if marca:
        condicoes.append("marca = %s")
        params.append(marca)

    vendedores = args.getlist('vendedor')
    if vendedores:
        placeholders = ','.join(['%s'] * len(vendedores))
        condicoes.append(f"vendedor IN ({placeholders})")
        params.extend(vendedores)

    where = ("WHERE " + " AND ".join(condicoes)) if condicoes else ""
    return where, params

def cache_key(rota, args):
    return rota + '?' + '&'.join(f'{k}={v}' for k, v in sorted(args.items()))

# ============================================================
# ROTAS
# ============================================================
@app.route('/')
def home():
    return jsonify({"status": "online", "mensagem": "API Horus funcionando!"})

@app.route('/api/filtros')
def filtros():
    key    = 'filtros'
    cached = cache_get(key)
    if cached: return jsonify(cached)

    anos      = consultar("SELECT DISTINCT ano FROM faturamento WHERE ano IS NOT NULL AND ano > 0 ORDER BY ano DESC")
    meses     = consultar("SELECT DISTINCT mes FROM faturamento WHERE mes IS NOT NULL AND mes > 0 ORDER BY mes")
    unidades  = consultar("SELECT DISTINCT unidade FROM faturamento WHERE unidade IS NOT NULL ORDER BY unidade")
    ufs       = consultar("SELECT DISTINCT uf FROM faturamento WHERE uf IS NOT NULL AND uf != '' ORDER BY uf")
    marcas    = consultar("SELECT DISTINCT marca FROM faturamento WHERE marca IS NOT NULL ORDER BY marca")
    tipos     = consultar("SELECT DISTINCT tipo_operacao FROM faturamento WHERE tipo_operacao IS NOT NULL ORDER BY tipo_operacao")
    vendedores = consultar("""
        SELECT DISTINCT vendedor
        FROM faturamento
        WHERE vendedor IS NOT NULL
        ORDER BY vendedor
    """)

    resultado = {
        "anos":       [r['ano']           for r in anos],
        "meses":      [r['mes']           for r in meses],
        "unidades":   [r['unidade']       for r in unidades],
        "ufs":        [r['uf']            for r in ufs],
        "marcas":     [r['marca']         for r in marcas],
        "tipos":      [r['tipo_operacao'] for r in tipos],
        "vendedores": [r['vendedor']      for r in vendedores],
    }
    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/kpis')
def kpis():
    key    = cache_key('kpis', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    resultado = consultar(f"""
        SELECT
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolução' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificação' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) /
                NULLIF(COUNT(CASE WHEN tipo_operacao = 'Venda' THEN 1 END), 0) AS NUMERIC), 2) AS ticket_medio,
            COUNT(DISTINCT cliente) AS total_clientes,
            COUNT(CASE WHEN tipo_operacao = 'Venda' THEN 1 END) AS qtd_vendas
        FROM faturamento {where}
    """, params)

    r = resultado[0] if resultado else {}
    cache_set(key, r)
    return jsonify(r)

@app.route('/api/faturamento-mensal')
def faturamento_mensal():
    key    = cache_key('faturamento-mensal', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    resultado = consultar(f"""
        SELECT ano, mes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolução' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificação' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes
        FROM faturamento {where}
        {'AND' if where else 'WHERE'} mes > 0
        GROUP BY ano, mes ORDER BY ano, mes
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-vendedores')
def top_vendedores():
    key    = cache_key('top-vendedores', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT vendedor,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes,
            COUNT(DISTINCT unidade) AS unidades,
            COUNT(*) AS qtd_vendas
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY vendedor ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-marca')
def faturamento_por_marca():
    key    = cache_key('faturamento-por-marca', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 15))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT marca,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY marca ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-regiao')
def faturamento_por_regiao():
    key    = cache_key('faturamento-por-regiao', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT regiao,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND regiao IS NOT NULL AND regiao != ''
        GROUP BY regiao ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-unidade')
def faturamento_por_unidade():
    key    = cache_key('faturamento-por-unidade', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT unidade,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolução' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificação' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} unidade IS NOT NULL
        GROUP BY unidade ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-produtos')
def top_produtos():
    key    = cache_key('top-produtos', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT produto, marca,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY produto, marca ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-uf')
def faturamento_por_uf():
    key    = cache_key('faturamento-por-uf', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT uf,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND uf IS NOT NULL AND uf != ''
        GROUP BY uf ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/carteira-vendedor')
def carteira_vendedor():
    key    = cache_key('carteira-vendedor', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    cod_vendedor = request.args.get('cod_vendedor', '')
    if cod_vendedor:
        resultado = consultar("""
            SELECT * FROM carteira WHERE cod_vendedor = %s ORDER BY cliente
        """, [cod_vendedor])
    else:
        resultado = consultar("""
            SELECT cod_vendedor, COUNT(*) as total_clientes
            FROM carteira GROUP BY cod_vendedor ORDER BY total_clientes DESC
        """)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/resumo-carteira')
def resumo_carteira():
    """
    Retorna total de clientes em carteira e margem média.
    Se filtrar por vendedor (nome), busca o cod_vendedor correspondente
    e retorna a carteira daquele vendedor.
    """
    key    = cache_key('resumo-carteira', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    vendedor     = request.args.get('vendedor', '')
    where, params = montar_filtros(request.args)
    and_or        = 'AND' if where else 'WHERE'

    margem = consultar(f"""
        SELECT ROUND(CAST(AVG(margem) AS NUMERIC), 2) AS margem_media
        FROM faturamento {where}
        {and_or} tipo_operacao = 'Venda'
        AND margem IS NOT NULL AND margem != 0
    """, params)
    margem_media = margem[0]['margem_media'] if margem else 0

    if vendedor:
        cods = consultar("""
            SELECT DISTINCT cod_vendedor FROM faturamento
            WHERE vendedor = %s AND cod_vendedor IS NOT NULL AND cod_vendedor != ''
        """, [vendedor])
        if cods:
            cod_list     = [c['cod_vendedor'] for c in cods]
            placeholders = ','.join(['%s'] * len(cod_list))
            carteira     = consultar(f"""
                SELECT COUNT(*) as total FROM carteira
                WHERE cod_vendedor IN ({placeholders})
            """, cod_list)
        else:
            carteira = [{'total': 0}]
    else:
        carteira = consultar("SELECT COUNT(*) as total FROM carteira")

    total_carteira = carteira[0]['total'] if carteira else 0
    resultado = {
        'total_carteira': total_carteira,
        'margem_media':   float(margem_media) if margem_media else 0,
    }
    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-clientes')
def top_clientes():
    key    = cache_key('top-clientes', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT cliente,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(*) AS qtd_vendas
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY cliente ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-cidade')
def faturamento_por_cidade():
    key    = cache_key('faturamento-por-cidade', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 15))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT cidade, uf,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND cidade IS NOT NULL AND cidade != ''
        GROUP BY cidade, uf ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/buscar-produtos')
def buscar_produtos():
    termo = request.args.get('q', '').strip()
    if not termo or len(termo) < 2:
        return jsonify([])

    resultado = consultar("""
        SELECT DISTINCT produto, cod_produto, marca
        FROM faturamento
        WHERE (LOWER(produto) LIKE LOWER(%s) OR LOWER(cod_produto) LIKE LOWER(%s))
        AND produto IS NOT NULL AND produto != ''
        ORDER BY produto
        LIMIT 30
    """, [f'%{termo}%', f'%{termo}%'])
    return jsonify(resultado)

@app.route('/api/top-produtos-filtrado')
def top_produtos_filtrado():
    key    = cache_key('top-produtos-filtrado', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    produtos = request.args.getlist('produtos')
    and_or   = 'AND' if where else 'WHERE'

    if produtos:
        placeholders = ','.join(['%s'] * len(produtos))
        resultado = consultar(f"""
            SELECT produto, marca,
                ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
                ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade,
                COUNT(DISTINCT cliente) AS clientes
            FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
            AND produto IN ({placeholders})
            GROUP BY produto, marca ORDER BY faturamento DESC
        """, params + produtos)
    else:
        limite    = int(request.args.get('limite', 20))
        resultado = consultar(f"""
            SELECT produto, marca,
                ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
                ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade,
                COUNT(DISTINCT cliente) AS clientes
            FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
            GROUP BY produto, marca ORDER BY faturamento DESC LIMIT %s
        """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/vendedores-por-produto')
def vendedores_por_produto():
    produtos = request.args.getlist('produtos')
    if not produtos:
        return jsonify([])

    placeholders = ','.join(['%s'] * len(produtos))
    resultado = consultar(f"""
        SELECT DISTINCT vendedor
        FROM faturamento
        WHERE produto IN ({placeholders})
        AND vendedor IS NOT NULL AND vendedor != ''
        AND tipo_operacao = 'Venda'
        ORDER BY vendedor
    """, produtos)
    return jsonify([r['vendedor'] for r in resultado])

# ============================================================
# SHELF LIFE — CONTROLE DE ACESSO
# ============================================================
# Lista de e-mails autorizados a acessar o Shelf Life
# Para adicionar ou remover alguém, edite esta lista
EMAILS_AUTORIZADOS_SL = [
    'comercial2@reforpan.com.br',
'comercial3@esdel.com.br'
    # adicione outros e-mails aqui
]

@app.route('/api/shelflife/verificar-acesso', methods=['POST'])
def shelflife_verificar_acesso():
    data  = request.get_json(force=True)
    email = str(data.get('email', '')).strip().lower()
    autorizado = email in [e.lower() for e in EMAILS_AUTORIZADOS_SL]
    return jsonify({'autorizado': autorizado, 'email': email})

# ============================================================
# SHELF LIFE
# ============================================================
from datetime import date as _date

@app.route('/api/shelflife/upload', methods=['POST'])
def shelflife_upload():
    data     = request.get_json()
    semana   = data.get('semana')
    unidade  = data.get('unidade')
    produtos = data.get('produtos', [])

    if not produtos:
        return jsonify({'erro': 'Nenhum produto enviado'}), 400

    conn   = get_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM shelflife WHERE semana = %s AND unidade = %s", [semana, unidade])

    hoje     = _date.today()
    inseridos = 0

    for p in produtos:
        validade = p.get('validade', '')
        try:
            from datetime import datetime
            val_date = datetime.strptime(str(validade)[:10], '%Y-%m-%d').date()
            dias     = (val_date - hoje).days
        except:
            dias = 999

        if dias <= 30:   status = 'CRITICO'
        elif dias <= 60: status = 'ATENCAO'
        else:            status = 'OK'

        nome  = str(p.get('produto', ''))
        is_sl = nome.upper().startswith('SL') or nome.upper().startswith('SL.')

        cursor.execute("""
            INSERT INTO shelflife (
                semana, unidade, cod_produto, cod_sl, produto, marca,
                quantidade_log, validade, dias_vencimento,
                vence_em, status_logistica, status, is_sl
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [
            semana, unidade,
            p.get('cod_produto'), p.get('cod_sl'),
            nome, p.get('marca'),
            p.get('quantidade'), validade[:10] if validade and len(str(validade)) >= 10 else None,
            dias, p.get('vence_em'), p.get('status_logistica'),
            status, is_sl
        ])
        inseridos += 1

    conn.commit(); cursor.close(); conn.close()
    return jsonify({'inseridos': inseridos, 'semana': semana, 'unidade': unidade})

@app.route('/api/shelflife/listar')
def shelflife_listar():
    import datetime as _dt
    semana  = request.args.get('semana')
    unidade = request.args.get('unidade')
    status  = request.args.get('status')
    where   = []; params = []

    if semana:
        where.append("semana = %s"); params.append(semana)
    else:
        where.append("semana = (SELECT MAX(semana) FROM shelflife)")

    if unidade:
        where.append("unidade = %s"); params.append(unidade)

    # Filtro de status é aplicado APÓS recalcular, então não filtra no SQL por status/is_sl aqui
    where_str = "WHERE " + " AND ".join(where) if where else ""
    resultado = consultar(f"SELECT * FROM shelflife {where_str} ORDER BY validade ASC NULLS LAST", params)

    # Recalcula dias_vencimento e status com base na data de HOJE
    hoje = _dt.date.today()
    status_map = []
    for row in resultado:
        val = row.get('validade')
        if val:
            try:
                val_date = _dt.date.fromisoformat(str(val)[:10])
                dias = (val_date - hoje).days
            except Exception:
                dias = row.get('dias_vencimento', 999)
        else:
            dias = row.get('dias_vencimento', 999)

        if dias <= 30:   novo_status = 'CRITICO'
        elif dias <= 60: novo_status = 'ATENCAO'
        else:            novo_status = 'OK'

        row['dias_vencimento'] = dias
        if not row.get('is_sl'):
            row['status'] = novo_status
        status_map.append(row)

    # Aplica filtro de status depois do recalculate (se solicitado)
    if status == 'SL':
        status_map = [r for r in status_map if r.get('is_sl')]
    elif status:
        status_map = [r for r in status_map if r.get('status') == status and not r.get('is_sl')]

    # Reordena por dias_vencimento atualizado
    status_map.sort(key=lambda r: r.get('dias_vencimento', 9999))

    return jsonify(status_map)

@app.route('/api/shelflife/semanas')
def shelflife_semanas():
    resultado = consultar("""
        SELECT DISTINCT semana, unidade, COUNT(*) as total
        FROM shelflife GROUP BY semana, unidade ORDER BY semana DESC
    """)
    return jsonify(resultado)

@app.route('/api/shelflife/atualizar', methods=['POST'])
def shelflife_atualizar():
    data    = request.get_json()
    id_prod = data.get('id')
    conn    = get_conn()
    cursor  = conn.cursor()

    cursor.execute("""
        SELECT semana, unidade, cod_produto, produto, quantidade_log,
               quantidade_atual, venda_3meses, venda_mes,
               data_inconsistencia, obs_logistica, obs_gerais, acao, vendedor
        FROM shelflife WHERE id = %s
    """, [id_prod])
    row = cursor.fetchone()

    # data_inconsistencia:
    #   None        → preserva o valor atual do banco (campo não foi tocado)
    #   '' (vazio)  → apaga intencionalmente (SET NULL)
    #   'yyyy-mm-dd'→ atualiza para a nova data
    _data_inc_raw = data.get('data_inconsistencia')
    if _data_inc_raw is None:
        # Não toca no banco — usa COALESCE para manter o valor existente
        _data_inc_sql = 'data_inconsistencia'   # será usado via literal no SQL
        _data_inc_params = []
    elif _data_inc_raw == '':
        # Apagar intencionalmente
        _data_inc_sql = 'NULL'
        _data_inc_params = []
    else:
        # Atualizar para nova data
        _data_inc_sql = '%s::date'
        _data_inc_params = [_data_inc_raw]

    cursor.execute(f"""
        UPDATE shelflife SET
            quantidade_atual     = %s,
            venda_3meses         = %s,
            venda_mes            = %s,
            data_inconsistencia  = {_data_inc_sql},
            obs_logistica        = %s,
            obs_gerais           = %s,
            acao                 = %s,
            vendedor             = %s,
            resolvido            = %s,
            updated_at           = NOW()
        WHERE id = %s
    """, [
        data.get('quantidade_atual'),
        data.get('venda_3meses'),
        data.get('venda_mes'),
        *_data_inc_params,
        data.get('obs_logistica'),
        data.get('obs_gerais'),
        data.get('acao'),
        data.get('vendedor'),
        data.get('resolvido', False),
        id_prod
    ])

    if row:
        campos = [
            ('quantidade_atual',    row[5],              data.get('quantidade_atual'),           'Qtde Atual'),
            ('venda_3meses',        row[6],              data.get('venda_3meses'),               'Venda 3 Meses'),
            ('venda_mes',           row[7],              data.get('venda_mes'),                  'Venda Mensal'),
            ('data_inconsistencia', str(row[8]) if row[8] else '', data.get('data_inconsistencia') or '', 'Data Inconsistencia'),
            ('obs_logistica',       row[9],              data.get('obs_logistica'),              'Obs. Logistica'),
            ('obs_gerais',          row[10],             data.get('obs_gerais'),                 'Obs. Gerais'),
            ('acao',                row[11],             data.get('acao'),                       'Acao'),
            ('vendedor',            row[12],             data.get('vendedor'),                   'Vendedor'),
        ]
        alteracoes = []
        for campo, antes, depois, label in campos:
            antes_str  = str(antes  or '').strip()
            depois_str = str(depois or '').strip()
            if antes_str != depois_str:
                alteracoes.append(f"{label}: [{antes_str or '—'}] → [{depois_str or '—'}]")

        if alteracoes:
            cursor.execute("""
                INSERT INTO shelflife_historico (
                    shelflife_id, semana, unidade, cod_produto, produto,
                    quantidade_log, quantidade_atual, venda_3meses, venda_mes,
                    acao, vendedor, obs_logistica, obs_gerais,
                    data_inconsistencia, usuario
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, [
                id_prod, row[0], row[1], row[2], row[3],
                row[4],
                data.get('quantidade_atual'),
                data.get('venda_3meses'),
                data.get('venda_mes'),
                data.get('acao'),
                data.get('vendedor'),
                '|'.join(alteracoes),
                data.get('obs_gerais'),
                data.get('data_inconsistencia') or None,
                data.get('usuario', 'admin')
            ])

    conn.commit(); cursor.close(); conn.close()
    return jsonify({'ok': True, 'alteracoes': alteracoes if row else []})

@app.route('/api/shelflife/excluir', methods=['POST'])
def shelflife_excluir():
    data    = request.get_json()
    semana  = data.get('semana')
    unidade = data.get('unidade')
    conn    = get_conn()
    cursor  = conn.cursor()

    if unidade:
        cursor.execute("DELETE FROM shelflife WHERE semana = %s AND unidade = %s", [semana, unidade])
    else:
        cursor.execute("DELETE FROM shelflife WHERE semana = %s", [semana])

    deleted = cursor.rowcount
    conn.commit(); cursor.close(); conn.close()
    return jsonify({'excluidos': deleted})

# ============================================================
# SHELF LIFE — EXPORTAR EXCEL FORMATADO
# ============================================================
@app.route('/api/shelflife/exportar', methods=['POST'])
def shelflife_exportar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data   = request.get_json(force=True)
    linhas = data.get('linhas', [])
    if not linhas:
        return jsonify({'erro': 'Nenhum dado para exportar'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'Shelf Life'

    HEADER_BG = '1A1A2E'; HEADER_FG = 'FFFFFF'
    ROW_ODD   = 'F7F8FA'; ROW_EVEN  = 'FFFFFF'
    STATUS_COLORS = {
        'SL':      ('7B2FBE', 'FFFFFF'),
        'CRITICO': ('FF4D4D', 'FFFFFF'),
        'ATENCAO': ('FF9800', 'FFFFFF'),
        'OK':      ('4CAF50', 'FFFFFF'),
        'NORMAL':  ('4CAF50', 'FFFFFF'),
        'ZERADO':  ('9E9E9E', 'FFFFFF'),
    }

    thin     = Side(style='thin',   color='CCCCCC')
    thin_d   = Side(style='thin',   color='0D0D1F')
    medium_d = Side(style='medium', color='0D0D1F')
    b_data   = Border(top=thin,   bottom=thin,     left=thin,   right=thin)
    b_hdr    = Border(top=thin_d, bottom=medium_d, left=thin_d, right=thin_d)

    col_widths = [10, 12, 12, 36, 18, 8, 14, 12, 13, 14, 15, 14, 19, 26, 26, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NUM_COLS    = {7, 8, 10, 11, 12}
    CENTER_COLS = {1, 2, 3, 6, 9, 10, 13, 16}

    # Colunas que devem ser numéricas (1-based): Qtde Log, Qtde Atual, Dias, V3M, VMes
    COLS_INT   = {7, 8, 10, 11, 12}  # mesmas de NUM_COLS

    def to_num(v):
        """Converte string para int ou float; retorna None se vazio/inválido."""
        if v is None or str(v).strip() == '':
            return None
        try:
            f = float(str(v).replace(',', '.'))
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return v   # mantém original se não for número

    for r_idx, row in enumerate(linhas, 1):
        ws.row_dimensions[r_idx].height = 22 if r_idx == 1 else 18
        is_hdr = (r_idx == 1)
        row_bg = ROW_ODD if r_idx % 2 == 0 else ROW_EVEN

        for c_idx, value in enumerate(row, 1):
            # Converte colunas numéricas para número real (evita bandeira verde do Excel)
            if not is_hdr and c_idx in COLS_INT:
                value = to_num(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if is_hdr:
                cell.font      = Font(bold=True, color=HEADER_FG, name='Segoe UI', size=10)
                cell.fill      = PatternFill('solid', fgColor=HEADER_BG)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = b_hdr
            else:
                cell.border = b_data
                if c_idx == 1:
                    sk = str(value or '').upper().strip()
                    if sk in STATUS_COLORS:
                        bg, fg = STATUS_COLORS[sk]
                        cell.fill = PatternFill('solid', fgColor=bg)
                        cell.font = Font(bold=True, color=fg, name='Segoe UI', size=9)
                    else:
                        cell.fill = PatternFill('solid', fgColor=row_bg)
                        cell.font = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in NUM_COLS:
                    cell.fill          = PatternFill('solid', fgColor=row_bg)
                    cell.font          = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'
                elif c_idx in CENTER_COLS:
                    cell.fill      = PatternFill('solid', fgColor=row_bg)
                    cell.font      = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill      = PatternFill('solid', fgColor=row_bg)
                    cell.font      = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Horus-ShelfLife.xlsx'
    )

@app.route('/api/pivot-clientes')
def pivot_clientes_novo():
    """
    Tabela pivot de clientes:
    - Sem produto: todos da carteira (LEFT JOIN com faturamento)
    - Com produto: só quem comprou (INNER JOIN)
    """
    vendedores = request.args.getlist('vendedor')
    produtos   = request.args.getlist('produtos')
    anos       = request.args.getlist('ano')
    conn       = get_conn()
    cursor     = conn.cursor()

    if produtos:
        prod_ph     = ','.join(['%s'] * len(produtos))
        vend_filter = ''
        params      = list(produtos)
        if vendedores:
            vend_ph     = ','.join(['%s'] * len(vendedores))
            vend_filter = f'AND f.vendedor IN ({vend_ph})'
            params     += vendedores
        ano_filter = ''
        if anos:
            ano_ph     = ','.join(['%s'] * len(anos))
            ano_filter = f'AND f.ano IN ({ano_ph})'
            params    += [int(a) for a in anos]

        cursor.execute(f"""
            SELECT
                f.cliente, f.cod_cliente, f.vendedor,
                f.ano, f.mes,
                ROUND(SUM(CASE WHEN f.tipo_operacao='Venda'    THEN f.valor_nf ELSE 0 END)::NUMERIC,2) AS faturamento,
                ROUND(SUM(CASE WHEN f.tipo_operacao='Devolucao' THEN f.valor_nf ELSE 0 END)::NUMERIC,2) AS devolucoes
            FROM faturamento f
            WHERE f.produto IN ({prod_ph})
            AND f.tipo_operacao IN ('Venda','Devolucao')
            AND f.mes > 0
            {vend_filter}
            {ano_filter}
            GROUP BY f.cliente, f.cod_cliente, f.vendedor, f.ano, f.mes
            ORDER BY f.cliente, f.ano, f.mes
        """, params)

    else:
        vend_filter = ''
        params      = []
        if vendedores:
            vend_ph     = ','.join(['%s'] * len(vendedores))
            vend_filter = f'WHERE c.cod_vendedor IN ({vend_ph})'
            params     += vendedores
        ano_filter = ''
        ano_params = []
        if anos:
            ano_ph     = ','.join(['%s'] * len(anos))
            ano_filter = f'AND f.ano IN ({ano_ph})'
            ano_params = [int(a) for a in anos]

        cursor.execute(f"""
            SELECT
                c.cliente, c.cod_cliente,
                v.vendedor,
                f.ano, f.mes,
                ROUND(COALESCE(SUM(CASE WHEN f.tipo_operacao='Venda'    THEN f.valor_nf ELSE 0 END),0)::NUMERIC,2) AS faturamento,
                ROUND(COALESCE(SUM(CASE WHEN f.tipo_operacao='Devolucao' THEN f.valor_nf ELSE 0 END),0)::NUMERIC,2) AS devolucoes
            FROM carteira c
            LEFT JOIN (
                SELECT DISTINCT cod_vendedor, vendedor FROM faturamento
            ) v ON v.cod_vendedor = c.cod_vendedor
            LEFT JOIN faturamento f
                ON f.cod_cliente = c.cod_cliente
                AND f.tipo_operacao IN ('Venda','Devolucao')
                AND f.mes > 0
                {ano_filter}
            {vend_filter}
            GROUP BY c.cliente, c.cod_cliente, v.vendedor, f.ano, f.mes
            ORDER BY c.cliente, f.ano, f.mes
        """, params + ano_params)

    rows      = cursor.fetchall()
    cols      = [desc[0] for desc in cursor.description]
    resultado = [dict(zip(cols, row)) for row in rows]
    cursor.close()
    conn.close()
    return jsonify(resultado)

@app.route('/api/shelflife/historico')
def shelflife_historico():
    shelflife_id = request.args.get('shelflife_id')
    cod_produto  = request.args.get('cod_produto')
    where        = []; params = []

    if shelflife_id:
        where.append("shelflife_id = %s"); params.append(shelflife_id)
    if cod_produto:
        where.append("cod_produto = %s");  params.append(cod_produto)

    where_str = "WHERE " + " AND ".join(where) if where else ""
    resultado = consultar(f"""
        SELECT * FROM shelflife_historico {where_str}
        ORDER BY created_at DESC LIMIT 50
    """, params)
    return jsonify(resultado)

# Limpa cache (útil após atualizar dados)
@app.route('/api/cache/clear', methods=['POST'])
def limpar_cache():
    cache_clear()
    return jsonify({"status": "cache limpo!"})

# Ping — mantém o servidor acordado
@app.route('/ping')
def ping():
    return jsonify({"status": "pong", "uptime": "ok"})

if __name__ == '__main__':
    print("🚀 API Horus iniciando...")
    app.run(debug=False, host='0.0.0.0', port=5000)
